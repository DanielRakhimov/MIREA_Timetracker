import sqlite3
import datetime as dt
import openpyxl
import sys
from io import StringIO

import unittest
import os

from openpyxl.reader.excel import load_workbook


class Application:
    def __init__(self, database):
        self.conn = sqlite3.connect(database)
        self.c = self.conn.cursor()
        self.c.execute('''CREATE TABLE IF NOT EXISTS events 
                          (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                           name TEXT, 
                           datetime DATETIME, 
                           description TEXT,
                           notification_sent BOOLEAN DEFAULT 0)''')
        self.conn.commit()

    def add_event(self, name, datetime, description):
        if datetime < dt.datetime.now():
            print("Ошибка: нельзя добавить мероприятие с прошедшей датой и временем")
            return
        self.c.execute("INSERT INTO events (name, datetime, description) VALUES (?, ?, ?)", (name, datetime, description))
        self.conn.commit()

    def delete_event(self, event_id):
        # Проверка наличия мероприятия по указанному ID
        self.c.execute("SELECT * FROM events WHERE id = ?", (event_id,))
        event = self.c.fetchone()

        if not event:
            print("Мероприятие не найдено")
            return

        # Удаление мероприятия
        self.c.execute("DELETE FROM events WHERE id = ?", (event_id,))
        self.conn.commit()
        print("Мероприятие успешно удалено")

    def get_upcoming_events(self):
        now = dt.datetime.now()
        start_of_day = dt.datetime.combine(now.date(), dt.time())
        self.c.execute("SELECT * FROM events WHERE datetime >= ? AND datetime < ? AND notification_sent = 0",
                       (now, start_of_day + dt.timedelta(days=1)))
        upcoming_events = self.c.fetchall()

        if len(upcoming_events) == 0:
            print("Запланированных событий на сегодня нет")

        return upcoming_events

    def send_notification(self):
        events = self.get_upcoming_events()
        for event in events:
            event_datetime = str(event[2])[:-3]
            message = f"Напоминаем, что мероприятие \"{event[1]}\" начнется {event_datetime}. {event[3]}"
            print(message)
            self.c.execute("UPDATE events SET notification_sent = 1 WHERE id = ?", (event[0],))
            self.conn.commit()

    def export_past_events(self):
        past_events = []
        now = dt.datetime.now()
        self.c.execute("SELECT id, name, datetime, description FROM events WHERE datetime < ?", (now,))
        past_events = self.c.fetchall()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["ID", "Название", "Дата и время", "Описание"])

        for event in past_events:
            worksheet.append(event)

        workbook.save('past_events.xlsx')

    def close_connection(self):
        self.conn.close()

'''
# Пример использования
app = Application('events.db')

# Добавление мероприятий
app.add_event("Встреча с клиентом", dt.datetime(2024, 1, 12, 9, 0), "Встреча в офисе")
app.add_event("Презентация", dt.datetime(2024, 2, 22, 21, 50), "Презентация нового продукта")

# Получение списка предстоящих мероприятий
upcoming_events = app.get_upcoming_events()
for event in upcoming_events:
    print(f"{event[0]} {event[1]} - {event[2]}")

# Удаление мероприятия по ID
app.delete_event(19)

# Отправка уведомления о старте мероприятия
app.send_notification()

# Выгрузка всех прошедших мероприятий в файл XLSX
app.export_past_events()

# Закрытие соединения с базой данных
app.close_connection()
'''
class TestApplication(unittest.TestCase):
    def setUp(self):
        self.db_path = 'test.db'  # Используем временную базу данных в памяти
        self.app = Application(self.db_path)

    def test_add_event(self):
        # Входные данные
        event_name = "Встреча с клиентом"
        event_date = dt.datetime(2024, 3, 12, 19, 40)
        event_date_str = "2024-03-12 19:40:00"
        event_description = "Встреча в офисе"

        # Выполняем тест
        self.app.add_event(event_name, event_date, event_description)

        # Получаем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM events WHERE name = ?", (event_name,))
        result = cursor.fetchone()


        # Проверяем ожидаемый результат
        self.assertIsNotNone(result)
        self.assertEqual(result[1], event_name)
        self.assertEqual(result[2], event_date_str)
        self.assertEqual(result[3], event_description)
        self.assertEqual(result[4], 0)  # Поле notification_sent должно быть равно 0
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()

    def test_delete_event(self):
        # Записи в БД
        events_data = [
            (1, "Презентация", dt.datetime(2024, 3, 12, 18, 20), "Презентация нового продукта", 1),
            (2, "Встреча", dt.datetime(2024, 3, 12, 18, 39), "Встреча в офисе", 0)
        ]


        for event in events_data:
            self.app.add_event(event[1], event[2], event[3])


        # Удаляем мероприятие с ID = 1
        self.app.delete_event(1)

        # Получаем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM events")
        result = cursor.fetchall()
        conn.close()

        # Проверяем ожидаемый результат
        expected_result = [
            (2, "Встреча", "2024-03-12 18:39:00", "Встреча в офисе", 0)
        ]

        self.assertEqual(result, expected_result)
        # Удаляем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()

    def test_delete_nonexistent_event(self):
        # Записи в БД
        events_data = [
            (1, "Презентация", dt.datetime(2024, 3, 12, 18, 20), "Презентация нового продукта", 1),
            (2, "Встреча", dt.datetime(2024, 3, 12, 18, 39), "Встреча в офисе", 0)
        ]

        for event in events_data:
            self.app.add_event(event[1], event[2], event[3])
            # Перенаправляем вывод из консоли в строку
        original_stdout = sys.stdout
        s = StringIO()
        sys.stdout = s

        # Вызываем функцию удаления
        self.app.delete_event(3)

        # Возвращаем обратно стандартный вывод
        sys.stdout = original_stdout

        # Получаем строку, выведенную в консоль
        console_output = s.getvalue().strip()

        # Проверяем, что строка содержит сообщение об ошибке
        self.assertIn("Мероприятие не найдено", console_output)

        # Получаем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM events")
        result = cursor.fetchall()
        conn.close()

        # Проверяем, что записи в БД не изменились
        expected_result = [
            (1, "Презентация", "2024-03-12 18:20:00", "Презентация нового продукта", 0),
            (2, "Встреча", "2024-03-12 18:39:00", "Встреча в офисе", 0)
        ]

        self.assertEqual(result, expected_result)
        # Удаляем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()

    def test_send_notification(self):
        # Запись в БД
        event_data = (1, "Презентация", dt.datetime(2024, 3, 9, 18, 39), "Презентация нового продукта")
        self.app.add_event(event_data[1],event_data[2], event_data[3])
        # Перенаправляем вывод из консоли в строку
        original_stdout = sys.stdout
        s = StringIO()
        sys.stdout = s

        # Вызываем функцию отправки уведомления
        self.app.send_notification()

        # Получаем строку, выведенную в консоль
        console_output = s.getvalue().strip()

        # Проверяем, что строка содержит правильное уведомление
        expected_output = 'Напоминаем, что мероприятие "Презентация" начнется 2024-03-09 18:39. Презентация нового продукта'
        self.assertIn(expected_output, console_output)

        # Возвращаем обратно стандартный вывод
        sys.stdout = original_stdout

        # Получаем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM events")
        result = cursor.fetchall()
        conn.close()

        # Проверяем, что поле notification_sent обновлено
        expected_result = [(1, "Презентация", "2024-03-09 18:39:00", "Презентация нового продукта", 1)]
        self.assertEqual(result, expected_result)

        # Удаляем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()

    def test_get_upcoming_events(self):
        # Запись в БД
        event_data_1 = (4, "Презентация", dt.datetime(2024, 2, 9, 18, 20), "Презентация нового продукта", 1)
        event_data_2 = (5, "Встреча", dt.datetime(2024, 3, 9, 18, 39), "Встреча в офисе", 0)

        # Заполняем базу данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS events 
                          (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                           name TEXT, 
                           datetime DATETIME, 
                           description TEXT,
                           notification_sent BOOLEAN DEFAULT 0)''')

        cursor.execute("INSERT INTO events VALUES (?, ?, ?, ?, ?)", event_data_1)
        cursor.execute("INSERT INTO events VALUES (?, ?, ?, ?, ?)", event_data_2)
        conn.commit()
        conn.close()

        # Вызываем функцию получения предстоящих мероприятий
        upcoming_events = self.app.get_upcoming_events()

        # Проверяем, что возвращается ожидаемый результат
        expected_result = [(5, "Встреча", "2024-03-09 18:39:00", "Встреча в офисе", 0)]
        self.assertEqual(upcoming_events, expected_result)

        # Удаляем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()

    def test_export_past_events(self):
        # Запись в БД
        event_data_1 = (1, "Презентация", dt.datetime(2024, 2, 12, 18, 20), "Презентация нового продукта", 1)
        event_data_2 = (2, "Встреча", dt.datetime(2024, 3, 12, 18, 39), "Встреча в офисе", 0)

        # Заполняем базу данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS events 
                          (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                           name TEXT, 
                           datetime DATETIME, 
                           description TEXT,
                           notification_sent BOOLEAN DEFAULT 0)''')

        cursor.execute("INSERT INTO events VALUES (?, ?, ?, ?, ?)", event_data_1)
        cursor.execute("INSERT INTO events VALUES (?, ?, ?, ?, ?)", event_data_2)
        conn.commit()
        conn.close()

        # Вызываем функцию выгрузки прошедших мероприятий в файл
        self.app.export_past_events()

        # Проверяем, что файл past_events.xlsx создан
        file_path = 'past_events.xlsx'
        self.assertTrue(os.path.isfile(file_path))

        # Загружаем файл и проверяем его содержимое
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        # Проверяем, что заголовки столбцов присутствуют
        headers = worksheet[1]
        headers = [cell.value for cell in headers]
        self.assertEqual(headers, ['ID', 'Название', 'Дата и время', 'Описание'])

        # Проверяем, что в файле есть ожидаемые записи
        expected_data = [1, 'Презентация', '2024-02-12 18:20:00', 'Презентация нового продукта']
        actual_row = [cell.value for cell in worksheet[2]]
        for i in range(0, 4):
            self.assertEqual(actual_row[i], expected_data[i])

        # Удаляем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()

    def test_add_past_event(self):
        # Записи в БД
        event = (1, "Презентация", dt.datetime(2024, 2, 12, 18, 20), "Презентация нового продукта", 1)

        # Перенаправляем вывод из консоли в строку
        original_stdout = sys.stdout
        s = StringIO()
        sys.stdout = s

        self.app.add_event(event[1], event[2], event[3])

        # Возвращаем обратно стандартный вывод
        sys.stdout = original_stdout

        # Получаем строку, выведенную в консоль
        console_output = s.getvalue().strip()

        # Проверяем, что строка содержит сообщение об ошибке
        self.assertIn("Ошибка: нельзя добавить мероприятие с прошедшей датой и временем", console_output)

        # Получаем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM events")
        result = cursor.fetchall()
        conn.close()

        # Проверяем, что записи в БД не изменились
        expected_result = []

        self.assertEqual(result, expected_result)
        # Удаляем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()

    def test_get_upcoming_events_when_no_events(self):
        # Перенаправляем вывод из консоли в строку
        original_stdout = sys.stdout
        s = StringIO()
        sys.stdout = s

        # Вызываем функцию получения предстоящих мероприятий
        upcoming_events = self.app.get_upcoming_events()

        # Возвращаем обратно стандартный вывод
        sys.stdout = original_stdout

        # Получаем строку, выведенную в консоль
        console_output = s.getvalue().strip()

        # Проверяем, что строка содержит сообщение об отсутствии мероприятий
        self.assertIn("Запланированных событий на сегодня нет", console_output)

        # Проверяем, что возвращается ожидаемый результат
        expected_result = []
        self.assertEqual(upcoming_events, expected_result)

        # Удаляем информацию из базы данных
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS events", ())
        conn.close()
    def tearDown(self):
        self.app.close_connection()

if __name__ == '__main__':
    unittest.main()
